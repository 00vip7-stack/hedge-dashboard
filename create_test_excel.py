#!/usr/bin/env python3
"""
테스트용 샘플 Excel 파일 생성
더존 ERP 형식 및 기본 형식으로 샘플 파일을 만듭니다.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import os

def create_dojeon_format():
    """더존 ERP 형식 샘플 파일"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "거래내역"
    
    # 헤더 설정 (더존 형식)
    headers = ['거래처명', '통화', '금액', '거래일자', '거래은행', '거래유형']
    ws.append(headers)
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 샘플 데이터
    sample_data = [
        ['ABC Import Co.', 'USD', 150000, datetime.now().date(), 'Citibank', 'Import'],
        ['XYZ Trading Ltd.', 'EUR', 250000, (datetime.now() - timedelta(days=3)).date(), 'Deutsche Bank', 'Export'],
        ['Global Supply Corp.', 'GBP', 80000, (datetime.now() - timedelta(days=7)).date(), 'HSBC', 'Import'],
        ['Pacific Trade Inc.', 'JPY', 5000000, (datetime.now() - timedelta(days=10)).date(), 'Mizuho Bank', 'Export'],
        ['European Partners', 'CHF', 120000, (datetime.now() - timedelta(days=14)).date(), 'UBS', 'Import'],
        ['Asian Commerce Ltd.', 'AUD', 200000, (datetime.now() - timedelta(days=1)).date(), 'ANZ Bank', 'Export'],
    ]
    
    for row_data in sample_data:
        ws.append(row_data)
    
    # 컬럼 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 금액 컬럼 포맷 (숫자)
    for row in ws.iter_rows(min_row=2, max_row=len(sample_data) + 1, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '#,##0'
    
    filename = 'sample_dojeon_format.xlsx'
    wb.save(filename)
    print(f"✅ 생성됨: {filename}")
    return filename

def create_basic_format():
    """기본 형식 샘플 파일"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # 헤더 설정 (기본 형식)
    headers = ['Counterparty', 'Currency', 'Amount', 'Date', 'Bank', 'Type']
    ws.append(headers)
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 샘플 데이터
    sample_data = [
        ['Supplier A', 'USD', 75000, datetime.now().date(), 'Bank A', 'Payment'],
        ['Vendor B', 'EUR', 125000, (datetime.now() - timedelta(days=5)).date(), 'Bank B', 'Invoice'],
        ['Partner C', 'GBP', 45000, (datetime.now() - timedelta(days=2)).date(), 'Bank C', 'Deposit'],
    ]
    
    for row_data in sample_data:
        ws.append(row_data)
    
    # 컬럼 너비
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 금액 포맷
    for row in ws.iter_rows(min_row=2, max_row=len(sample_data) + 1, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = '#,##0'
    
    filename = 'sample_basic_format.xlsx'
    wb.save(filename)
    print(f"✅ 생성됨: {filename}")
    return filename

def create_complex_format():
    """복잡한 헤더명을 가진 샘플 파일 (컬럼 매핑 테스트용)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "거래내역"
    
    # 복잡한 헤더 (더존 ERP 형식과 유사)
    headers = ['거래처 이름', '외화 코드', '외화 금액', '거래 일자', '적요', '용도']
    ws.append(headers)
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 샘플 데이터
    sample_data = [
        ['국제 무역 회사', 'USD', 500000, '2026-02-04', '원자재 수입', '제조'],
        ['현지 납품업체', 'EUR', 300000, '2026-02-03', '부품 구매', '조립'],
        ['해외 파트너사', 'GBP', 150000, '2026-01-28', '서비스 계약금', '유지보수'],
    ]
    
    for row_data in sample_data:
        ws.append(row_data)
    
    # 컬럼 너비
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    filename = 'sample_complex_format.xlsx'
    wb.save(filename)
    print(f"✅ 생성됨: {filename}")
    return filename

if __name__ == '__main__':
    print("🔨 테스트용 Excel 샘플 파일 생성 중...\n")
    
    try:
        create_dojeon_format()
        create_basic_format()
        create_complex_format()
        
        print("\n✨ 모든 샘플 파일이 생성되었습니다!")
        print("\n📋 생성된 파일:")
        print("  1. sample_dojeon_format.xlsx - 더존 ERP 형식")
        print("  2. sample_basic_format.xlsx - 기본 형식")
        print("  3. sample_complex_format.xlsx - 복잡한 헤더명 형식")
        print("\n💡 이 파일들을 HedgeFreedom에 업로드하여 테스트하세요.")
        
    except ImportError:
        print("❌ openpyxl 라이브러리가 필요합니다.")
        print("설치: pip install openpyxl")
    except Exception as e:
        print(f"❌ 파일 생성 중 오류: {e}")
