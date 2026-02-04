"""
Excel 파일 파싱 및 데이터 익명화 모듈
HedgeFreedom - powered by hedgeOn Engine
"""

import openpyxl
from datetime import datetime, timedelta
import re
import hashlib
import random


class ExcelParser:
    """엑셀 파일 파싱 클래스"""
    
    def __init__(self, file_path=None, file_stream=None):
        """
        Args:
            file_path: 파일 경로
            file_stream: 파일 스트림 (BytesIO 객체)
        """
        if file_path:
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
        elif file_stream:
            self.workbook = openpyxl.load_workbook(file_stream, data_only=True)
        else:
            raise ValueError("파일 경로 또는 파일 스트림이 필요합니다.")
        
        self.sheet = self.workbook.active
    
    def parse_trade_data(self):
        """
        거래 데이터 파싱
        
        예상 엑셀 구조:
        | 거래ID | 거래처명 | 통화 | 외화금액 | 결제예정일 | 수출/수입 | 헤지상태 |
        
        Returns:
            list: 거래 데이터 리스트
        """
        trades = []
        
        # 헤더 찾기 (첫 번째 행 또는 'ID', '거래처', '통화' 등 포함된 행)
        header_row = self._find_header_row()
        
        if header_row is None:
            # 헤더를 못 찾으면 1행을 헤더로 간주
            header_row = 1
        
        # 컬럼 매핑
        column_mapping = self._map_columns(header_row)
        
        # 데이터 행 파싱 (헤더 다음 행부터)
        for row_idx in range(header_row + 1, self.sheet.max_row + 1):
            row_data = self._parse_row(row_idx, column_mapping)
            
            if row_data and row_data.get('amount'):  # 금액이 있는 행만
                trades.append(row_data)
        
        return trades
    
    def _find_header_row(self):
        """헤더 행 찾기"""
        header_keywords = ['거래', 'ID', '통화', '금액', '날짜', 'trade', 'currency', 'amount']
        
        for row_idx in range(1, min(10, self.sheet.max_row + 1)):  # 최대 10행까지만 검색
            row_values = [str(cell.value).lower() if cell.value else '' 
                         for cell in self.sheet[row_idx]]
            
            # 키워드가 2개 이상 포함되면 헤더로 간주
            keyword_count = sum(1 for keyword in header_keywords 
                              if any(keyword.lower() in val for val in row_values))
            
            if keyword_count >= 2:
                return row_idx
        
        return None
    
    def _map_columns(self, header_row):
        """컬럼 매핑 생성"""
        mapping = {
            'trade_id': None,
            'counterparty': None,
            'currency': None,
            'amount': None,
            'settlement_date': None,
            'trade_type': None,
            'hedge_status': None
        }
        
        header_patterns = {
            'trade_id': ['거래id', 'tradeid', 'id', '번호'],
            'counterparty': ['거래처', '업체', '회사', 'counterparty', 'company', 'client'],
            'currency': ['통화', 'currency', 'curr'],
            'amount': ['금액', '외화금액', 'amount', 'value'],
            'settlement_date': ['결제', '만기', '예정일', 'settlement', 'maturity', 'date'],
            'trade_type': ['수출', '수입', '구분', 'type', 'export', 'import'],
            'hedge_status': ['헤지', 'hedge', '상태', 'status']
        }
        
        for col_idx in range(1, self.sheet.max_column + 1):
            cell_value = str(self.sheet.cell(header_row, col_idx).value or '').lower()
            
            for field, patterns in header_patterns.items():
                if any(pattern in cell_value for pattern in patterns):
                    mapping[field] = col_idx
                    break
        
        return mapping
    
    def _parse_row(self, row_idx, column_mapping):
        """행 데이터 파싱"""
        try:
            # 거래 ID
            trade_id = self._get_cell_value(row_idx, column_mapping['trade_id'])
            if not trade_id:
                trade_id = f"T{row_idx - 1}"
            
            # 거래처명
            counterparty = self._get_cell_value(row_idx, column_mapping['counterparty']) or '거래처'
            
            # 통화
            currency = self._get_cell_value(row_idx, column_mapping['currency']) or 'USD'
            currency = self._normalize_currency(currency)
            
            # 금액
            amount = self._get_cell_value(row_idx, column_mapping['amount'])
            if amount is None:
                return None
            amount = self._parse_number(amount)
            if amount <= 0:
                return None
            
            # 결제예정일
            settlement_date = self._get_cell_value(row_idx, column_mapping['settlement_date'])
            settlement_date = self._parse_date(settlement_date)
            
            # 거래 구분
            trade_type = self._get_cell_value(row_idx, column_mapping['trade_type']) or '수출'
            trade_type = self._normalize_trade_type(trade_type)
            
            # 헤지 상태
            hedge_status = self._get_cell_value(row_idx, column_mapping['hedge_status']) or '미헤지'
            hedge_status = self._normalize_hedge_status(hedge_status)
            
            # 원화환산액 계산
            exchange_rate = self._get_exchange_rate(currency)
            krw_amount = int(amount * exchange_rate)
            
            # D-Day 계산
            if settlement_date:
                days_until = (settlement_date - datetime.now()).days
            else:
                days_until = 30
            
            return {
                'id': str(trade_id),
                'counterparty': str(counterparty),
                'currency': currency,
                'amount': amount,
                'settlementDate': settlement_date.strftime('%Y-%m-%d') if settlement_date else None,
                'type': trade_type,
                'krwAmount': krw_amount,
                'daysUntil': days_until,
                'hedgeStatus': hedge_status
            }
            
        except Exception as e:
            print(f"행 {row_idx} 파싱 오류: {e}")
            return None
    
    def _get_cell_value(self, row_idx, col_idx):
        """셀 값 가져오기"""
        if col_idx is None:
            return None
        
        cell = self.sheet.cell(row_idx, col_idx)
        return cell.value
    
    def _parse_number(self, value):
        """숫자 파싱"""
        if isinstance(value, (int, float)):
            return float(value)
        
        if isinstance(value, str):
            # 쉼표, 공백 제거
            value = value.replace(',', '').replace(' ', '')
            try:
                return float(value)
            except:
                return 0
        
        return 0
    
    def _parse_date(self, value):
        """날짜 파싱"""
        if isinstance(value, datetime):
            return value
        
        if isinstance(value, str):
            # 다양한 날짜 형식 지원
            date_formats = ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%m/%d/%Y', '%d/%m/%Y']
            
            for fmt in date_formats:
                try:
                    return datetime.strptime(value, fmt)
                except:
                    continue
        
        # 파싱 실패 시 30일 후로 설정
        return datetime.now() + timedelta(days=30)
    
    def _normalize_currency(self, value):
        """통화 정규화"""
        value = str(value).upper().strip()
        
        currency_map = {
            'USD': 'USD', 'US': 'USD', 'DOLLAR': 'USD', '달러': 'USD',
            'EUR': 'EUR', 'EURO': 'EUR', '유로': 'EUR',
            'JPY': 'JPY', 'YEN': 'JPY', '엔': 'JPY',
            'CNY': 'CNY', 'RMB': 'CNY', '위안': 'CNY',
            'GBP': 'GBP', 'POUND': 'GBP', '파운드': 'GBP'
        }
        
        return currency_map.get(value, 'USD')
    
    def _normalize_trade_type(self, value):
        """거래 구분 정규화"""
        value = str(value).lower()
        
        if '수출' in value or 'export' in value:
            return '수출'
        elif '수입' in value or 'import' in value:
            return '수입'
        else:
            return '수출'
    
    def _normalize_hedge_status(self, value):
        """헤지 상태 정규화"""
        value = str(value).lower()
        
        if '전액' in value or 'full' in value or '완료' in value:
            return '전액헤지'
        elif '부분' in value or 'partial' in value:
            return '부분헤지'
        else:
            return '미헤지'
    
    def _get_exchange_rate(self, currency):
        """환율 가져오기 (Mock)"""
        rates = {
            'USD': 1350.0,
            'EUR': 1450.0,
            'JPY': 9.5,
            'CNY': 185.0,
            'GBP': 1700.0
        }
        return rates.get(currency, 1350.0)


class DataAnonymizer:
    """데이터 익명화 클래스"""
    
    def __init__(self, customer_id):
        self.customer_id = customer_id
        self.anonymization_map = {}
    
    def anonymize_trades(self, trades):
        """
        거래 데이터 익명화
        
        Args:
            trades: 거래 데이터 리스트
        
        Returns:
            list: 익명화된 거래 데이터
        """
        anonymized_trades = []
        
        for trade in trades:
            anonymized_trade = trade.copy()
            
            # 거래처명 익명화
            anonymized_trade['counterparty'] = self._anonymize_counterparty(trade['counterparty'])
            
            # 거래 ID 해시화 (선택적)
            # anonymized_trade['id'] = self._hash_value(trade['id'])
            
            anonymized_trades.append(anonymized_trade)
        
        return anonymized_trades
    
    def _anonymize_counterparty(self, counterparty):
        """거래처명 익명화"""
        # 이미 매핑된 경우
        if counterparty in self.anonymization_map:
            return self.anonymization_map[counterparty]
        
        # 개인정보 패턴 제거
        anonymized = self._remove_personal_info(counterparty)
        
        # 알파벳으로 매핑
        index = len(self.anonymization_map)
        if index < 26:
            anonymized_name = f"거래처{chr(65 + index)}"  # A, B, C, ...
        else:
            anonymized_name = f"거래처{index + 1}"
        
        self.anonymization_map[counterparty] = anonymized_name
        
        return anonymized_name
    
    def _remove_personal_info(self, text):
        """개인정보 패턴 제거"""
        if not text:
            return text
        
        text = str(text)
        
        # 이메일 제거
        text = re.sub(r'[\w\.-]+@[\w\.-]+\.\w+', '[이메일제거]', text)
        
        # 전화번호 제거
        text = re.sub(r'\d{2,3}-\d{3,4}-\d{4}', '[전화번호제거]', text)
        text = re.sub(r'\d{10,11}', '[전화번호제거]', text)
        
        # 주민등록번호 제거
        text = re.sub(r'\d{6}-\d{7}', '[주민번호제거]', text)
        
        # 사업자등록번호 제거
        text = re.sub(r'\d{3}-\d{2}-\d{5}', '[사업자번호제거]', text)
        
        return text
    
    def _hash_value(self, value):
        """값 해시화"""
        if not value:
            return value
        
        # SHA-256 해시 (고객 ID와 함께 해시하여 고객별로 다른 해시 생성)
        hash_input = f"{self.customer_id}:{value}"
        hash_value = hashlib.sha256(hash_input.encode()).hexdigest()
        
        return hash_value[:16]  # 16자리만 사용


def calculate_kpi(positions):
    """
    KPI 계산
    
    Args:
        positions: 포지션 데이터 리스트
    
    Returns:
        dict: KPI 데이터
    """
    if not positions:
        return {
            'totalExposure': 0,
            'hedgedAmount': 0,
            'currentHedgeRatio': 0,
            'targetHedgeRatio': 70,
            'gap': -70,
            'unhedgedAmount': 0
        }
    
    # 총 노출액
    total_exposure = sum(p['krwAmount'] for p in positions)
    
    # 헤지된 금액
    hedged_amount = sum(
        p['krwAmount'] for p in positions 
        if p['hedgeStatus'] in ['전액헤지', '부분헤지']
    )
    
    # 부분헤지는 50%만 계산 (간단한 가정)
    partial_hedge = sum(
        p['krwAmount'] * 0.5 for p in positions 
        if p['hedgeStatus'] == '부분헤지'
    )
    
    full_hedge = sum(
        p['krwAmount'] for p in positions 
        if p['hedgeStatus'] == '전액헤지'
    )
    
    actual_hedged = full_hedge + partial_hedge
    
    # 헤지 비율
    current_ratio = (actual_hedged / total_exposure * 100) if total_exposure > 0 else 0
    
    # 목표 비율 (기본값 70%)
    target_ratio = 70
    
    # 갭
    gap = current_ratio - target_ratio
    
    # 미헤지 금액
    unhedged_amount = total_exposure - actual_hedged
    
    return {
        'totalExposure': int(total_exposure),
        'hedgedAmount': int(actual_hedged),
        'currentHedgeRatio': round(current_ratio, 1),
        'targetHedgeRatio': target_ratio,
        'gap': round(gap, 1),
        'unhedgedAmount': int(unhedged_amount)
    }
