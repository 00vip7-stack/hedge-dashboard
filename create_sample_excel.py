"""
샘플 엑셀 데이터 생성기
테스트용 거래 데이터 엑셀 파일 생성
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
import random


def create_sample_excel(filename='sample_trades.xlsx'):
    """샘플 거래 데이터 엑셀 파일 생성"""
    
    # 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "거래데이터"
    
    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 작성
    headers = [
        "거래ID",
        "거래처명",
        "통화",
        "외화금액",
        "결제예정일",
        "수출/수입",
        "헤지상태"
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 샘플 데이터
    companies = [
        "ABC Trading Co.",
        "글로벌상사",
        "XYZ Corporation",
        "한국무역(주)",
        "International Exports Ltd.",
        "대한수출입",
        "Pacific Trade Inc.",
        "아시아무역",
        "Euro Partners GmbH",
        "동양상사"
    ]
    
    currencies = ['USD', 'EUR', 'JPY', 'CNY']
    trade_types = ['수출', '수입']
    hedge_statuses = ['미헤지', '부분헤지', '전액헤지']
    
    # 데이터 행 생성
    for i in range(20):
        row_idx = i + 2
        
        # 거래 ID
        ws.cell(row=row_idx, column=1).value = f"T{2024000 + i + 1}"
        
        # 거래처명
        ws.cell(row=row_idx, column=2).value = random.choice(companies)
        
        # 통화
        currency = random.choice(currencies)
        ws.cell(row=row_idx, column=3).value = currency
        
        # 외화금액
        if currency == 'JPY':
            amount = random.randint(1000000, 50000000)
        else:
            amount = random.randint(10000, 500000)
        ws.cell(row=row_idx, column=4).value = amount
        
        # 결제예정일
        days_ahead = random.randint(7, 180)
        settlement_date = datetime.now() + timedelta(days=days_ahead)
        ws.cell(row=row_idx, column=5).value = settlement_date.strftime('%Y-%m-%d')
        
        # 수출/수입
        ws.cell(row=row_idx, column=6).value = random.choice(trade_types)
        
        # 헤지상태
        ws.cell(row=row_idx, column=7).value = random.choice(hedge_statuses)
    
    # 컬럼 너비 조정
    column_widths = [12, 25, 10, 15, 15, 12, 12]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # 파일 저장
    wb.save(filename)
    print(f"✅ 샘플 엑셀 파일 생성 완료: {filename}")
    
    return filename


if __name__ == "__main__":
    # 샘플 파일 생성
    create_sample_excel('sample_trades.xlsx')
    
    # 다양한 케이스 테스트용 파일들
    print("\n추가 테스트 파일들:")
    
    # 케이스 1: 헤더가 없는 경우
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['T2024001', '테스트회사', 'USD', 50000, '2026-03-15', '수출', '미헤지'])
    ws.append(['T2024002', '샘플거래처', 'EUR', 30000, '2026-04-20', '수입', '부분헤지'])
    wb.save('sample_no_header.xlsx')
    print("  - sample_no_header.xlsx (헤더 없음)")
    
    # 케이스 2: 다양한 날짜 형식
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['거래ID', '거래처명', '통화', '금액', '날짜', '구분', '헤지'])
    ws.append(['T001', '회사A', 'USD', 100000, '2026/03/15', 'export', 'unhedged'])
    ws.append(['T002', '회사B', 'EUR', 50000, '2026.04.20', 'import', 'partial'])
    wb.save('sample_various_formats.xlsx')
    print("  - sample_various_formats.xlsx (다양한 형식)")
    
    print("\n모든 샘플 파일 생성 완료!")
